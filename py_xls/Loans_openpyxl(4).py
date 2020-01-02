import openpyxl
wb = openpyxl.load_workbook(filename='test(01).xlsx')
sheet = wb.get_sheet_by_name('Loans')
#print(sheet.cell(column=4, row=4).value)

import sqlite3
 
conn = sqlite3.connect("one03.db") 
cursor = conn.cursor()

z = 2
while sheet.cell(column=1, row=z).value:
       one1 = (sheet.cell(column=1, row=z).value)
       one2 = '"' + str(sheet.cell(column=2,row=z).value) + '"'
       one3 = sheet.cell(column=3, row=z).value
       one4 = sheet.cell(column=4, row=z).value
       one5 = sheet.cell(column=5, row=z).value
       # print('one1 = ', one1)

       f1 = "insert into Loans (LoanId, DisbursmentDate, UserType, Term, Amount)"
       f2 = " VALUES ( {0}, {1}, {2}, {3}, {4});".format(one1, one2, one3, one4, one5)
       #print('f2 = ',f2)
       f3 = f1+f2
       #print('f3 = ',f3)
       cursor.execute(f3)
       conn.commit()
       z = z + 1
       print('текущая позиция = ', str(z).center(5), '\r', end = '')


"""

CREATE TABLE Loans (                -- Loans	таблиця-довідник по кредитам
        LoanId INT(8) primary key,     -- (integer) LoanId	ідентифікатор кредиту
        DisbursmentDate DATE,       -- DisbursmentDate	дата видачі
        UserType bit,               -- UserType	ідентифікатор типу клієнта
        -- FOREIGN KEY для UserType (table UserTypes)
        -- связь с UserTypes
        Term TINYINT,               -- Term	термін, на який видали кредит
        Amount DECIMAL(5,2) );      -- Amount	сума кредиту


"""